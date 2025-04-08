from starlette.responses import FileResponse
from fastapi import FastAPI
from fastapi.responses import HTMLResponse
from fastapi import BackgroundTasks
import os
from pathlib import Path
import random
import pandas as pd
import datetime
from datetime import timedelta
import calendar
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import copy
from openpyxl.styles import PatternFill, Alignment
from pydantic import BaseModel, Field
from fastapi.middleware.cors import CORSMiddleware

# this class is responsible for the dates in general


app = FastAPI()
origins = [
    "http://127.0.0.1:3000",  # Allow Flask (front-end) to make requests
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],  # Allow all HTTP methods
    allow_headers=["*"], )  # Allow all headers




@app.get("/labnya", response_class=HTMLResponse)
async def read_root():
    html_path = Path("templates/hajj.html").read_text(encoding="utf-8")
    return HTMLResponse(content=html_path)
    
class Data(BaseModel):
    e_num: int
    holi: list = Field(default=[])  # Default holidays
    monthh: int = Field(default=0)
    vac: dict = Field(default={})  # Default vacation data





@app.get("/", response_class=HTMLResponse)
async def read_root():
    html_path = Path("templates/homepage.html").read_text()
    return HTMLResponse(content=html_path)



@app.get('/d', response_class=HTMLResponse)
async def dis():
    html_page = Path("templates/index.html").read_text()
    return HTMLResponse(content=html_page)

@app.post('/d')
async def get(data: Data, background_tasks: BackgroundTasks):
    global names_list, vacations, month_name, week_end
    if data.e_num <=20:
        data.e_num = 30
    class Tframe:
        def __init__(self, datee):
            self.date = datee
            global week_end, month_name
            week_end = []
            month_name = ''

        #  fills the variables with the dates from the next month
        def next_nmonth(self, m):  # fills the dates of the next 30 days
            global week_end, month_name
            days_of_the_month = calendar.monthrange(self.date.year, self.date.month)[1]
            days_till_end_month = days_of_the_month - self.date.day
            self.date = self.date + timedelta(days=days_till_end_month + 1)
            self.date = datetime.datetime(self.date.year - 1, 12, 1)
            self.date = self.date + relativedelta(months=m)
            days_of_the_month = calendar.monthrange(self.date.year, self.date.month)[1]
            month_name = self.date.strftime('%B')
            days_name = []
            print('this is before the i loop')
            print(self.date)
            print(week_end)

            for i in range(days_of_the_month):
                new_d = self.date + timedelta(days=i)
                main_keys_days.append(new_d.strftime("%A"))
                if new_d.strftime("%A") == "Friday" or new_d.strftime("%A") == "Saturday":
                    week_end.append(str(int(new_d.strftime("%d"))))
                days_name.append(new_d.strftime("%A"))
                new_d = str(int(new_d.strftime("%d")))
                new_d = f"{new_d}"
                main_dic.update({new_d: []})
                main_keys.append(new_d)

        def next_month(self):
            global week_end, month_name
            days_of_the_month = calendar.monthrange(self.date.year, self.date.month)[1]
            days_till_end_month = days_of_the_month - self.date.day
            self.date = self.date + timedelta(days=days_till_end_month + 1)
            days_of_the_month = calendar.monthrange(self.date.year, self.date.month)[1]
            month_name = self.date.strftime('%B')
            days_name = []

            for i in range(days_of_the_month):
                new_d = self.date + timedelta(days=i)
                main_keys_days.append(new_d.strftime("%A"))
                if new_d.strftime("%A") == "Friday" or new_d.strftime("%A") == "Saturday":
                    week_end.append(str(int(new_d.strftime("%d"))))
                days_name.append(new_d.strftime("%A"))
                new_d = str(int(new_d.strftime("%d")))
                new_d = f"{new_d}"
                main_dic.update({new_d: []})
                main_keys.append(new_d)

    # could add a function that it fills the date with any month on demand

    class Eframe:
        def __init__(self, main_di, main_di_v, name_list):
            global name, N, index, selection_counts, month_name, week_end, excel_file
            self.dic_values = main_di_v.copy()
            self.WK_dic = copy.deepcopy(main_di)
            self.N_dic = copy.deepcopy(main_di)
            self.PM_dic = copy.deepcopy(main_di)
            self.names = name_list.copy()

        def N(self):
            global name, N, index, selection_counts
            dic_vals = self.dic_values.copy()
            N_names_list = self.names.copy()

            index = 0  # Start from the first name

            random.shuffle(N_names_list)
            random.shuffle(dic_vals)
            random.shuffle(N_names_list)
            random.shuffle(N_names_list)

            selection_counts = {name: 0 for name in N_names_list}  # Initialize counts to zero

            for i in range(len(main_keys)):
                k = 0  # Counter to ensure four unique names are added per key
                while k < 4:
                    # Sort names by the number of times they've been assigned, ascending
                    sorted_names = sorted(selection_counts, key=selection_counts.get)

                    sort_len = len(sorted_names)
                    # Attempt to assign the least-used name

                    for nam in range(sort_len):
                        sorted_names = sorted(selection_counts, key=selection_counts.get)
                        name = sorted_names[nam]

                        # Check for collisions across `AM_dic` and `PM_dic`
                        if main_keys[i] in vacations[name]:
                            pass
                        elif name not in self.N_dic[main_keys[i]] and name not in self.N_dic[main_keys[i - 1]]:

                            self.N_dic[main_keys[i]].append(name)  # Add to AM_dic
                            selection_counts[name] += 1  # Increment the count for this name
                            k += 1  # Increment unique count for this key
                            break  # Exit inner loop to move to the next unique position

            N_names_list = self.names.copy()

            N = self.N_dic.copy()

            for N_e in range(len(main_keys)):
                empp = [""] * len(N_names_list)

                key = N[main_keys[N_e]]
                for i in range(len(N_names_list)):
                    # print(i)
                    tel = key

                    if N_names_list[i] in tel:
                        empp[i] = "N"
                N[main_keys[N_e]] = empp

        def PM(self):
            global name, PM, index
            dic_val = self.dic_values.copy()
            PM_names_list = self.names.copy()

            index = 0  # Start from the first name

            random.shuffle(PM_names_list)
            random.shuffle(dic_val)
            random.shuffle(PM_names_list)
            random.shuffle(PM_names_list)

            for i in range(len(main_keys)):
                k = 0  # Counter to ensure four unique names are added per key
                while k < 4:

                    # Sort names by the number of times they've been assigned, ascending
                    sorted_names = sorted(selection_counts, key=selection_counts.get)

                    sort_len = len(sorted_names)
                    # Attempt to assign the least-used name

                    for nam in range(sort_len):
                        sorted_names = sorted(selection_counts, key=selection_counts.get)
                        name = sorted_names[nam]

                        # Check for collisions across `AM_dic` and `PM_dic`
                        if main_keys[i] in vacations[name]:
                            pass
                        elif name not in self.PM_dic[main_keys[i]] and name not in self.N_dic[main_keys[i]] \
                                and name not in self.N_dic[main_keys[i - 1]]:
                            self.PM_dic[main_keys[i]].append(name)  # Add to AM_dic
                            selection_counts[name] += 1  # Increment the count for this name
                            k += 1  # Increment unique count for this key
                            break  # Exit inner loop to move to the next unique position

            PM_names_list = self.names.copy()

            PM = self.PM_dic.copy()
            for PM_e in range(len(main_keys)):
                empp = [""] * len(PM_names_list)

                key = PM[main_keys[PM_e]]
                for i in range(len(PM_names_list)):
                    tel = key

                    if PM_names_list[i] in tel:
                        empp[i] = "PM"
                PM[main_keys[PM_e]] = empp

        def WK(self):
            global name, WK, index, week_end
            dic_vals = self.dic_values.copy()
            WK_names_list = self.names.copy()

            index = 0  # Start from the first name
            random.shuffle(WK_names_list)
            random.shuffle(dic_vals)
            random.shuffle(WK_names_list)
            random.shuffle(WK_names_list)

            for i in range(len(week_end)):
                k = 0
                while k < 4:

                    sorted_names = sorted(selection_counts, key=selection_counts.get)

                    sort_len = len(sorted_names)

                    for nam in range(sort_len):
                        sorted_names = sorted(selection_counts, key=selection_counts.get)
                        name = sorted_names[nam]

                        # Check for collisions across `AM_dic` and `PM_dic`

                        if main_keys[i] in vacations[name]:
                            pass
                        elif name not in self.PM_dic[week_end[i]] \
                        and name not in self.N_dic[main_keys[int(week_end[i]) - 2]] \
                        and name not in self.WK_dic[week_end[i]] \
                        and name not in self.N_dic[main_keys[int(week_end[i]) - 1]]:
                            self.WK_dic[week_end[i]].append(name)  # Add to AM_dic
                            selection_counts[name] += 1  # Increment the count for this name
                            k += 1  # Increment unique count for this key
                            break  # Exit inner loop to move to the next unique position

            WK_names_list = self.names.copy()

            WK = self.WK_dic.copy()
            for WK_e in range(len(main_keys)):
                empp = [""] * len(WK_names_list)

                key = WK[main_keys[WK_e]]
                for i in range(len(WK_names_list)):
                    tel = key

                    if WK_names_list[i] in tel:
                        empp[i] = "AM"
                WK[main_keys[WK_e]] = empp

        def print(self):
            global excel_file
            data = {
                "names": names_list
            }

            def merge_dicts(dict1, dict2, dict3):
                merged_dict = {}

                for key in dict1.keys():
                    # Get the lists from all three dictionaries
                    list1 = dict1[key]
                    list2 = dict2[key]
                    list3 = dict3[key]

                    # Initialize the merged list
                    merged_list = []

                    # Loop through the items in all three lists
                    for item1, item2, item3 in zip(list1, list2, list3):
                        if item1 == item2 == item3:
                            # If all items are the same, keep that item
                            merged_list.append(item1)
                        elif ("N" in [item1, item2, item3] and "PM" in [item1, item2, item3]) or \
                                ("N" in [item1, item2, item3] and "AM" in [item1, item2, item3]) or \
                                ("PM" in [item1, item2, item3] and "AM" in [item1, item2, item3]):
                            # If any combination of AM, PM, and WK appears, mark as collision
                            merged_list.append("collision")
                        elif (item1 in ["N", "PM", "AM"] and item2 == "" and item3 == "") or \
                                (item2 in ["N", "PM", "AM"] and item1 == "" and item3 == "") or \
                                (item3 in ["N", "PM", "AM"] and item1 == "" and item2 == ""):
                            # If only one is AM, PM, or WK and the others are empty, keep the non-empty item
                            merged_list.append(item1 if item1 else item2 if item2 else item3)

                            pass
                        else:
                            # For any other cases, retain empty or other default values

                            merged_list.append("")
                    merged_dict[key] = merged_list

                if all(not lst for lst in vacations.values()):
                    return merged_dict
                else:
                    cleaned_vac = {key: value for key, value in vacations.items() if value}
                    diclist = list(cleaned_vac.keys())
                    for v in range(len(diclist)):
                        name = diclist[v]
                        index = names_list.index(int(name))

                        for x in range(len(cleaned_vac[name])):
                            merged_dict[cleaned_vac[name][x]][index] = 'H'

                    return merged_dict

            main_keys_days.insert(0, ' ')

            data.update(merge_dicts(PM, N, WK))
            df = pd.DataFrame(data)
            df.to_excel('Schedule.xlsx', index=False, engine='openpyxl')
            wb = load_workbook('Schedule.xlsx')  # Replace with your file name
            ws = wb.active
            ws.insert_rows(1)
            for col_num, value in enumerate(main_keys_days, start=1):  # start=1 to start from column A
                ws.cell(row=1, column=col_num, value=value[0])
            max_ro = ws.max_row
            max_co = ws.max_column
            dark_fill = PatternFill(start_color="909090", end_color="909090", fill_type="solid")
            for w in range(len(week_end)):
                for r in range(int(max_ro) - 2):
                    ws.cell(row=3 + r, column=1 + int(week_end[w])).fill = dark_fill
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:  # Only format non-empty cells
                        cell.alignment = Alignment(horizontal="center")
            excel_file = 'Schedule.xlsx'
            wb.save('Schedule.xlsx')





        def count_shifts(self):
            for wq in range(len(self.names)):
                namess = self.names[wq]
                s = 0
                for d in range(len(self.dic_values)):
                    rand = self.dic_values[d]
                    if namess in self.PM_dic[rand]:
                        s += 1
                    if namess in self.N_dic[rand]:
                        s += 1
                    if namess in self.WK_dic[rand]:
                        s += 1

                if s == 9 or s == 6:
                    print(f"{namess}={s}erororororororororooror")
                else:
                    print(f"{namess}={s}")

        def count_days_shifts(self):
            for ein in range(len(self.dic_values)):
                l = len(self.N_dic[self.dic_values[ein]]) + len(self.PM_dic[self.dic_values[ein]])
                ll = l + len(self.WK_dic[self.dic_values[ein]])

        def count_emps(self):
            namelist = self.names.copy()
            print(f"we have {len(namelist)} available employees")

    main_dic = {}  # each key is a date and contains the names of people who are working on these days
    main_keys = []  # a list of each key easier to handle
    main_keys_days = []  # names of the dates like m for monday


    date = datetime.datetime.now()



    names_list = list(range(1, data.e_num + 1))
    vacations = {key: [] for key in names_list}
    tot = len(data.vac.keys())
    k = list(data.vac.keys())
    for h in range(tot):
        # print(f"this is vac_k[h] {data.vac[k[h]]}")
        # print(f"this is k[h] {k[h]}")

        lis = []
        # print(data.vac[k[h]][0])
        for i in range(int(data.vac[k[h]][0]), int(data.vac[k[h]][1]) + 1):  # +1 to include the end number
            lis.append(str(i))

        vacations[k[h]] = lis


    T = Tframe(date)  # T will fill the necessary list to be able to distribute emps shifts
    if data.monthh == 0:
        T.next_month()
    else:
        T.next_nmonth(data.monthh)
    week_end.extend(data.holi)
    print(week_end)
    e = Eframe(main_dic, main_keys, names_list)

    e.N()
    e.PM()
    e.WK()

    e.print()
    file_path = excel_file
    # Path to the file
    background_tasks.add_task(os.unlink, file_path)
    return FileResponse(file_path,
                        media_type="application/octet-stream",
                        filename=f'{file_path}')

# remember use uvicorn {thjsfilename}:{fastapi varible} --reload
# remember use uvicorn web_api:app --reload

# if i dont have postman use the route /docs automaticly shows me what i need
