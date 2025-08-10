import copy
import random
from openpyxl.styles import PatternFill, Alignment
import pandas as pd
from openpyxl import load_workbook



class Eframe:
    def __init__(self, main_dic, main_keys, name_list, shift_limit, data,vacation, main_keys_days, week_end):
        global name, N, index, selection_counts, month_name, excel_file
        self.main_keys = main_keys.copy()
        self.WK_dic = copy.deepcopy(main_dic)
        self.N_dic = copy.deepcopy(main_dic)
        self.PM_dic = copy.deepcopy(main_dic)
        self.names = name_list.copy()
        self.shift_limit = shift_limit
        self.data = data
        self.vacation = vacation
        self.main_keys_days = main_keys_days
        self.week_end = week_end

    def N(self):
        global name, N, index, selection_counts
        dic_vals = self.main_keys_days.copy()
        N_names_list = self.names.copy()

        index = 0  # Start from the first name

        random.shuffle(N_names_list)
        random.shuffle(dic_vals)
        random.shuffle(N_names_list)
        random.shuffle(N_names_list)

        selection_counts = {name: 0 for name in N_names_list}  # Initialize counts to zero

        for i in range(len(self.main_keys)):
            k = 0  # Counter to ensure four unique names are added per key
            nt = 0  # while loop counter
            broke = False

            while k < self.shift_limit:
                nt += 1
                if nt > len(self.names):
                    print("broke N shift")
                    print(len(self.names))
                    print(nt)
                    broke = True
                    break

                # Sort names by the number of times they've been assigned, ascending
                sorted_names = sorted(selection_counts, key=selection_counts.get)

                sort_len = len(sorted_names)
                # Attempt to assign the least-used name

                for nam in range(sort_len):
                    sorted_names = sorted(selection_counts, key=selection_counts.get)
                    name = sorted_names[nam]

                    # Check for collisions across `AM_dic` and `PM_dic`
                    if self.main_keys[i] in self.vacation[name]:
                        pass
                    elif name not in self.N_dic[self.main_keys[i]] and name not in self.N_dic[self.main_keys[i - 1]]:
                        print(self.data.n_shift)
                        if self.data.n_shift == 1:
                            self.N_dic[self.main_keys[i]].append(name)  # Add to AM_dic
                            selection_counts[name] += 1  # Increment the count for this name
                        k += 1  # Increment unique count for this key
                        break  # Exit inner loop to move to the next unique position
            if broke:
                break
        N_names_list = self.names.copy()

        N = self.N_dic.copy()

        for N_e in range(len(self.main_keys)):
            empp = [""] * len(N_names_list)

            key = N[self.main_keys[N_e]]
            for i in range(len(N_names_list)):
                # print(i)
                tel = key

                if N_names_list[i] in tel:
                    empp[i] = "N"
            N[self.main_keys[N_e]] = empp

    def PM(self):
        global name, PM, index
        dic_val = self.main_keys.copy()
        PM_names_list = self.names.copy()

        index = 0  # Start from the first name

        random.shuffle(PM_names_list)
        random.shuffle(dic_val)
        random.shuffle(PM_names_list)
        random.shuffle(PM_names_list)


        for i in range(len(self.main_keys)):
            k = 0  # Counter to ensure four unique names are added per key
            PM_c = 0  # while loop counter

            while k < self.shift_limit:
                if PM_c >> len(self.main_keys):
                    print("broke PM shift")
                    break
                PM_c += 1

                # Sort names by the number of times they've been assigned, ascending
                sorted_names = sorted(selection_counts, key=selection_counts.get)

                sort_len = len(sorted_names)
                # Attempt to assign the least-used name

                for nam in range(sort_len):
                    sorted_names = sorted(selection_counts, key=selection_counts.get)
                    name = sorted_names[nam]

                    # Check for collisions across `AM_dic` and `PM_dic`
                    if self.main_keys[i] in self.vacation[name]:
                        pass
                    elif name not in self.PM_dic[self.main_keys[i]] and name not in self.N_dic[self.main_keys[i]] \
                            and name not in self.N_dic[self.main_keys[i - 1]]:
                        self.PM_dic[self.main_keys[i]].append(name)  # Add to AM_dic
                        selection_counts[name] += 1  # Increment the count for this name
                        k += 1  # Increment unique count for this key
                        break  # Exit inner loop to move to the next unique position

        PM_names_list = self.names.copy()

        PM = self.PM_dic.copy()
        for PM_e in range(len(self.main_keys)):
            empp = [""] * len(PM_names_list)

            key = PM[self.main_keys[PM_e]]
            for i in range(len(PM_names_list)):
                tel = key

                if PM_names_list[i] in tel:
                    empp[i] = "PM"
            PM[self.main_keys[PM_e]] = empp

    def WK(self):
        global name, WK, index
        dic_vals = self.main_keys.copy()
        WK_names_list = self.names.copy()

        index = 0  # Start from the first name
        random.shuffle(WK_names_list)
        random.shuffle(dic_vals)
        random.shuffle(WK_names_list)
        random.shuffle(WK_names_list)

        for i in range(len(self.week_end)):
            k = 0
            WK_c = 0  # while loop counter
            while k < self.shift_limit:
                if WK_c >> len(self.names):
                    print("broke PM shift")
                    break
                WK_c += 1

                sorted_names = sorted(selection_counts, key=selection_counts.get)

                sort_len = len(sorted_names)

                for nam in range(sort_len):
                    sorted_names = sorted(selection_counts, key=selection_counts.get)
                    name = sorted_names[nam]

                    # Check for collisions across `AM_dic` and `PM_dic`

                    if self.main_keys[i] in self.vacation[name]:
                        pass
                    elif name not in self.PM_dic[self.week_end[i]] \
                            and name not in self.N_dic[self.main_keys[int(self.week_end[i]) - 2]] \
                            and name not in self.WK_dic[self.week_end[i]] \
                            and name not in self.N_dic[self.main_keys[int(self.week_end[i]) - 1]]:
                        self.WK_dic[self.week_end[i]].append(name)  # Add to AM_dic
                        selection_counts[name] += 1  # Increment the count for this name
                        k += 1  # Increment unique count for this key
                        break  # Exit inner loop to move to the next unique position

        WK_names_list = self.names.copy()

        WK = self.WK_dic.copy()
        for WK_e in range(len(self.main_keys)):
            empp = [""] * len(WK_names_list)

            key = WK[self.main_keys[WK_e]]
            for i in range(len(WK_names_list)):
                tel = key

                if WK_names_list[i] in tel:
                    empp[i] = "AM"
            WK[self.main_keys[WK_e]] = empp

    def print(self):
        global excel_file
        dataa = {
            "names": self.names
        }

        def merge_dicts(dict1, dict2, dict3):
            merged_dict = {}

            for key in dict1.keys():
                # Get the lists from all three dictionaries
                list1 = dict1[key]
                list2 = dict2[key]
                list3 = dict3[key]

                # Initialize the merged list
                merged_list = []

                # Loop through the items in all three lists
                for item1, item2, item3 in zip(list1, list2, list3):
                    if item1 == item2 == item3:
                        # If all items are the same, keep that item
                        merged_list.append(item1)
                    elif ("N" in [item1, item2, item3] and "PM" in [item1, item2, item3]) or \
                            ("N" in [item1, item2, item3] and "AM" in [item1, item2, item3]) or \
                            ("PM" in [item1, item2, item3] and "AM" in [item1, item2, item3]):
                        # If any combination of AM, PM, and WK appears, mark as collision
                        merged_list.append("collision")
                    elif (item1 in ["N", "PM", "AM"] and item2 == "" and item3 == "") or \
                            (item2 in ["N", "PM", "AM"] and item1 == "" and item3 == "") or \
                            (item3 in ["N", "PM", "AM"] and item1 == "" and item2 == ""):
                        # If only one is AM, PM, or WK and the others are empty, keep the non-empty item
                        merged_list.append(item1 if item1 else item2 if item2 else item3)

                        pass
                    else:
                        # For any other cases, retain empty or other default values

                        merged_list.append("")
                merged_dict[key] = merged_list

            if all(not lst for lst in self.vacation.values()):
                return merged_dict
            else:
                cleaned_vac = {key: value for key, value in self.vacation.items() if value}
                diclist = list(cleaned_vac.keys())
                for v in range(len(diclist)):
                    name = diclist[v]
                    index = self.names.index(int(name))
                    print(f"this the index = {index}")

                    for x in range(len(cleaned_vac[name])):
                        merged_dict[cleaned_vac[name][x]][index] = 'H'
                print(merged_dict)
                return merged_dict

        self.main_keys_days.insert(0, ' ')

        dataa.update(merge_dicts(PM, N, WK))
        print(self.N_dic)
        print(self.PM_dic)
        print(self.WK_dic)
        for key, value in dataa.items():
            print(f"{key}: {value}")
            print(f"{key}: {len(value)}")
        print(dataa)
        df = pd.DataFrame(dataa)
        df.to_excel('Schedule.xlsx', index=False, engine='openpyxl')
        wb = load_workbook('Schedule.xlsx')  # Replace with your file name
        ws = wb.active
        ws.insert_rows(1)
        for col_num, value in enumerate(self.main_keys_days, start=1):  # start=1 to start from column A
            ws.cell(row=1, column=col_num, value=value[0])
        max_ro = ws.max_row
        max_co = ws.max_column
        dark_fill = PatternFill(start_color="909090", end_color="909090", fill_type="solid")
        for w in range(len(self.week_end)):
            for r in range(int(max_ro) - 2):
                ws.cell(row=3 + r, column=1 + int(self.week_end[w])).fill = dark_fill
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:  # Only format non-empty cells
                    cell.alignment = Alignment(horizontal="center")
        excel_file = 'schedule.xlsx'
        wb.save('schedule.xlsx')
        return excel_file

    def count_shifts(self):
        for wq in range(len(self.names)):
            namess = self.names[wq]
            s = 0
            for d in range(len(self.main_keys)):
                rand = self.main_keys[d]
                if namess in self.PM_dic[rand]:
                    s += 1
                if namess in self.N_dic[rand]:
                    s += 1
                if namess in self.WK_dic[rand]:
                    s += 1

            if s == 9 or s == 6:
                print(f"{namess}={s}erororororororororooror")
            else:
                print(f"{namess}={s}")

    def count_days_shifts(self):
        for ein in range(len(self.main_keys)):
            l = len(self.N_dic[self.main_keys[ein]]) + len(self.PM_dic[self.main_keys[ein]])
            ll = l + len(self.WK_dic[self.main_keys[ein]])

    def count_emps(self):
        namelist = self.names.copy()
        print(f"we have {len(namelist)} available employees")
